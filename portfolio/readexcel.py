# -*- coding: utf-8 -*-
"""
Created on Sun Jun 11 13:33:51 2017

@author: Stuart
"""

import openpyxl, pprint

def read(filename):
    print('Opening workbook...')
    wb = openpyxl.load_workbook(filename, read_only=True)
    sheet = wb.get_sheet_by_name('Portfolio')
    stockData = {}

    assert (sheet['C1'].value == 'Symbol'), "Symbol is not in expected column"
    assert (sheet['H1'].value == 'Curr.'), "Currency is not in expected column"
    assert (sheet['L1'].value == 'Type'), "Stock Exchange/Type is not in expected column"
    assert (sheet['I1'].value == 'Current Price'), "Current Price is not in expected column"
    assert (sheet['K1'].value == 'Date'), "Date is not in expected column"

    print('Reading rows...')
    for row in range(2, sheet.max_row + 1):
        # Each row in the spreadsheet has data for one census tract.
        symbol  = sheet['C' + str(row)].value
        currency = sheet['H' + str(row)].value
        exchange = sheet['L' + str(row)].value
        price = sheet['I' + str(row)].value
        date = sheet['K' + str(row)].value
        
        # Make sure the key for this symbol exists.
        #stockData.setdefault(symbol, {})
        # Make sure the key for this county in this state exists.
        if symbol not in stockData and symbol is not None:
            stockData[symbol] = {'currency': currency, 'exchange': exchange, 'price': price, 'date': date}
            #stockData[symbol].setdefault(symbol, {'currency': 0, 'pop': 0})
            
    return stockData

import os
def write(filepath, stockData, exchange_rate):
    print('Opening ' + filepath + ' for writing...')
    filename, file_extension = os.path.splitext(filepath)

    # if the file extension is xlsm then make sure we keep vba
    if file_extension == '.xlsm':
        wb = openpyxl.load_workbook(filepath, read_only=False, keep_vba=True)
    elif file_extension == '.xlsx':
         wb = openpyxl.load_workbook(filepath, read_only=False, keep_vba=False)
    else:
        raise Exception('Unsupported file type')
       
    sheet = wb.get_sheet_by_name('Portfolio')

    # get todays date
    from datetime import date
    today = date.today()

    # update rows with the new stock data
    print('Writing rows...')
    for row in range(2, sheet.max_row + 1):
        # Each row in the spreadsheet has data for one stock symbol
        symbol  = sheet['C' + str(row)].value
        if symbol is None:
            continue
        
        exchange = stockData[symbol]['exchange']
        
        if exchange != 'manual':
            print(symbol)
            if stockData[symbol]['price'] is None:
                print('Price not changed!')
            else:
                sheet['I' + str(row)] = stockData[symbol]['price']
                sheet['K' + str(row)] = today # TODO use actual stock retrival date eg. stockData[symbol]['date']

    if exchange_rate is not None:
        sheet = wb.get_sheet_by_name('constants')
        sheet['B1'].value = exchange_rate
    else:
        print('Exchange rate not updated!')
    
    wb.save('backup' + file_extension)
    wb.save(filepath)
 
            
if __name__ == "__main__":
    # execute only if run as a script
    stockData = read('book1.xlsx')
    print(stockData)
    stockData['AXP']['price'] = 9999
    write('book2.xlsx', stockData,3.5)



